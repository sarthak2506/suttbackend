import openpyxl
from openpyxl import load_workbook
import json
from datetime import datetime

def parse_mess_menu(excel_path, output_json_path):
    wb = load_workbook(excel_path)
    sheet = wb.active

    days_of_week = ['SATURDAY', 'SUNDAY', 'MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY']
    menu_data = {}

    for col in range(1, 16):
        date_cell = sheet.cell(row=2, column=col)
        date_value = date_cell.value

        if not date_value:
            continue

        if isinstance(date_value, datetime):
            date_str = date_value.strftime('%d-%b-%y').upper()
        else:
            date_str = str(date_value).strip()

        breakfast = []
        lunch = []
        dinner = []
        current_meal = None

        for row in sheet.iter_rows(min_row=3):
            cell = row[col - 1]
            value = cell.value

            if value in ['BREAKFAST', 'LUNCH', 'DINNER']:
                current_meal = value.title()
            elif value in days_of_week:
                current_meal = None
            else:
                if current_meal:
                    if value is not None:
                        stripped_value = str(value).strip()
                        if stripped_value and stripped_value != '********':
                            if current_meal == 'Breakfast':
                                breakfast.append(stripped_value)
                            elif current_meal == 'Lunch':
                                lunch.append(stripped_value)
                            elif current_meal == 'Dinner':
                                dinner.append(stripped_value)

        menu_data[date_str] = {
            'Breakfast': breakfast,
            'Lunch': lunch,
            'Dinner': dinner
        }

    with open(output_json_path, 'w', encoding='utf-8') as json_file:
        json.dump(menu_data, json_file, indent=2, ensure_ascii=False)

if __name__ == '__main__':
    input_excel = r'C:\Users\abc\Downloads\messmenu.xlsx'
    output_json = r'C:\Users\abc\Downloads\menu.json'
    parse_mess_menu(input_excel, output_json)